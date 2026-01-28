"""
Excel Export Module
Exporteert simulatieresultaten naar Excel formaat
"""

import sys
import os
from typing import Dict
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from models import Project
from berekeningen import SimulatieResultaat


def genereer_excel_export(
    bestandsnaam: str,
    project: Project,
    resultaten: Dict[str, SimulatieResultaat]
):
    """
    Genereer een Excel bestand met alle simulatiedata.
    
    Args:
        bestandsnaam: Pad naar output Excel bestand
        project: Project met gebouwgegevens
        resultaten: Simulatieresultaten per trap
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback: genereer CSV
        _genereer_csv_export(bestandsnaam, project, resultaten)
        return
    
    wb = openpyxl.Workbook()
    
    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2874A6', end_color='2874A6', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # === Sheet 1: Projectgegevens ===
    ws = wb.active
    ws.title = "Projectgegevens"
    
    ws['A1'] = "Evacuatie Doorstroom Capaciteitsberekening"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    project_data = [
        ["", ""],
        ["Projectnaam:", project.projectnaam],
        ["Scenario:", project.omschrijving],
        ["Projectnummer:", project.projectnummer],
        ["Gebruiker:", project.gebruiker],
        ["Datum berekeningen:", project.datum_berekeningen.strftime("%d-%m-%Y")],
        ["", ""],
        ["GEBOUWCONFIGURATIE", ""],
        ["Aantal bouwlagen:", project.aantal_bouwlagen],
        ["Laagste verdieping:", project.laagste_verdieping],
        ["Aantal trappen:", project.aantal_trappen],
        ["Totaal personen:", project.totaal_personen],
    ]
    
    for row_idx, row_data in enumerate(project_data, start=2):
        ws.cell(row=row_idx, column=1, value=row_data[0]).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=row_data[1])
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    
    # === Sheet 2: Samenvatting ===
    ws_sam = wb.create_sheet("Samenvatting")
    
    headers = ["Trap", "Personen", "Ontruimingstijd [min]", "Buiten na tijd", "Percentage"]
    for col, header in enumerate(headers, start=1):
        cell = ws_sam.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    for row_idx, (naam, res) in enumerate(resultaten.items(), start=2):
        ws_sam.cell(row=row_idx, column=1, value=naam).border = thin_border
        ws_sam.cell(row=row_idx, column=2, value=res.totaal_personen).border = thin_border
        ws_sam.cell(row=row_idx, column=3, value=res.ontruimingstijd_min).border = thin_border
        ws_sam.cell(row=row_idx, column=4, value=res.eindstand_buiten).border = thin_border
        
        percentage = (res.eindstand_buiten / res.totaal_personen * 100) if res.totaal_personen > 0 else 0
        cell = ws_sam.cell(row=row_idx, column=5, value=f"{percentage:.1f}%")
        cell.border = thin_border
        cell.alignment = center_align
    
    for col in range(1, 6):
        ws_sam.column_dimensions[get_column_letter(col)].width = 20
    
    # === Sheets per trap: Gedetailleerde resultaten ===
    for naam, res in resultaten.items():
        # Maak sheet naam veilig (max 31 karakters, geen speciale tekens)
        sheet_naam = naam.replace(" ", "_")[:31]
        ws_trap = wb.create_sheet(sheet_naam)
        
        # Headers
        headers = [
            "Tijdstap", "Tijd [min]", "Naar buiten", "Cumulatief buiten",
            "In trap", "Op verdiepingen", "In portalen"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws_trap.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # Data
        for row_idx, ts in enumerate(res.tijdstappen, start=2):
            ws_trap.cell(row=row_idx, column=1, value=ts.tijdstap).border = thin_border
            ws_trap.cell(row=row_idx, column=2, value=ts.tijd_min).border = thin_border
            ws_trap.cell(row=row_idx, column=3, value=round(ts.naar_buiten, 2)).border = thin_border
            ws_trap.cell(row=row_idx, column=4, value=round(ts.cumulatief_buiten, 2)).border = thin_border
            ws_trap.cell(row=row_idx, column=5, value=round(ts.totaal_in_trap, 2)).border = thin_border
            ws_trap.cell(row=row_idx, column=6, value=round(ts.totaal_op_verdiepingen, 2)).border = thin_border
            ws_trap.cell(row=row_idx, column=7, value=round(ts.totaal_in_portalen, 2)).border = thin_border
        
        # Kolom breedtes
        for col in range(1, 8):
            ws_trap.column_dimensions[get_column_letter(col)].width = 18
    
    # === Sheet: Personen per verdieping ===
    ws_pers = wb.create_sheet("Personen_per_verdieping")
    
    headers = ["Verdieping", "Aantal personen"]
    for col, header in enumerate(headers, start=1):
        cell = ws_pers.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    for row_idx, verd in enumerate(project.verdiepingen, start=2):
        ws_pers.cell(row=row_idx, column=1, value=verd.nummer).border = thin_border
        ws_pers.cell(row=row_idx, column=2, value=verd.aantal_personen).border = thin_border
    
    # Totaal rij
    totaal_row = len(project.verdiepingen) + 2
    ws_pers.cell(row=totaal_row, column=1, value="TOTAAL").font = Font(bold=True)
    ws_pers.cell(row=totaal_row, column=2, value=project.totaal_personen).font = Font(bold=True)
    
    ws_pers.column_dimensions['A'].width = 15
    ws_pers.column_dimensions['B'].width = 20
    
    # Opslaan
    wb.save(bestandsnaam)


def _genereer_csv_export(
    bestandsnaam: str,
    project: Project,
    resultaten: Dict[str, SimulatieResultaat]
):
    """Fallback: genereer CSV bestanden als openpyxl niet beschikbaar is"""
    import csv
    
    # Vervang .xlsx met _samenvatting.csv
    base_naam = bestandsnaam.rsplit('.', 1)[0] if '.' in bestandsnaam else bestandsnaam
    
    # Samenvatting CSV
    with open(f"{base_naam}_samenvatting.csv", 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, delimiter=';')
        
        writer.writerow(["Evacuatie Doorstroom Capaciteitsberekening"])
        writer.writerow([])
        writer.writerow(["Projectnaam", project.projectnaam])
        writer.writerow(["Scenario", project.omschrijving])
        writer.writerow(["Totaal personen", project.totaal_personen])
        writer.writerow([])
        writer.writerow(["Trap", "Personen", "Ontruimingstijd", "Buiten na tijd", "Percentage"])
        
        for naam, res in resultaten.items():
            percentage = (res.eindstand_buiten / res.totaal_personen * 100) if res.totaal_personen > 0 else 0
            writer.writerow([
                naam, res.totaal_personen, res.ontruimingstijd_min,
                f"{res.eindstand_buiten:.0f}", f"{percentage:.1f}%"
            ])
    
    # Per trap een CSV
    for naam, res in resultaten.items():
        trap_naam = naam.replace(" ", "_")
        with open(f"{base_naam}_{trap_naam}.csv", 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f, delimiter=';')
            
            writer.writerow(["Tijdstap", "Tijd [min]", "Naar buiten", "Cumulatief", "In trap", "Op verdiepingen"])
            
            for ts in res.tijdstappen:
                writer.writerow([
                    ts.tijdstap, ts.tijd_min,
                    f"{ts.naar_buiten:.2f}",
                    f"{ts.cumulatief_buiten:.2f}",
                    f"{ts.totaal_in_trap:.2f}",
                    f"{ts.totaal_op_verdiepingen:.2f}"
                ])
